#Embedded System Engineering 2021148007 기나혜
#기말 프로젝트: 소셜미디어의 개인정보(계정) 저장 프로그램
#목적: 각 SNS의 아이디와 패스워드를 한 엑셀파일로 저장하고, 콘솔 화면에서 검색, 추가, 수정 기능을 구현한다.
#본 프로그램은 variable, function, main interface로 구성되어있다.


import openpyxl
#파이썬에서 엑셀파일을 읽고, 쓸 수 있게 하는 모듈. 본 프로젝트의 핵심이다.
#openpyxl외에도 엑셀 파일을 다루는 모듈은 pandas, pyexcel등이 있다.
#특히 pandas는 표 형식으로 된 데이터를 다루기에 유용하여,
#본 프로젝트의 초기 제작계획은 pandas와 openpyxl을 병행하여 작업하려 하였으나,
#openpyxl 모듈만으로도 제작 의도를 구현할 수 있을 듯 하여 배제하였다.


############# Variable Code ################
filepath = 'saving_id_pw_file.xlsx'             
#데이터를 저장하는 엑셀파일명. 소스코드와 같은 폴더 내에 있을 경우 '파일명.확장자'만 적는다.
#다른 폴더에 있을 경우 절대경로를 할당하며, 파일명 앞에 r을 붙여 raw file name으로 작업한다.

datafile = openpyxl.load_workbook(filepath)
#엑셀파일 저장

sheet = datafile['Sheet1']
#datafile내 'Sheet1'명의 시트를 저장


sns_column = []
set_sns_column = []
id_column = []
pw_column = []
#리스트 초기화





############# Function Code ################
def offset():
    #기본 리스트 저장용 offset함수 -> 프로그램 시작, 엑셀파일 load 직후 offset() 설정
    #iter_rows() function : iterate each row
    #sheet의 행을 반복하여 각 행의 데이터에 access 가능하게 하는 함수이다.
    #cell : sheet의 셀 단위
    #cell.value : 특정 셀 내의 데이터를 가리킨다. 자료형은 string default.
    
    #1열의 SNS 데이터 access. append함수로 해당 list에 저장
    for row in sheet.iter_rows(min_col =1, max_col = 1):
        for cell in row:
            sns_column.append(cell.value)

    #2열의 ID 데이터 access
    for row in sheet.iter_rows(min_col =2, max_col = 2):
        for cell in row:
            id_column.append(cell.value)

    #3열의 PW 데이터 access
    for row in sheet.iter_rows(min_col =3, max_col = 3):
        for cell in row:
            pw_column.append(cell.value)

    #return 값 list형 sns_column, main code에서 사용
    return sns_column






def home_search(sns):
    #SNS, 플랫폼 search function
    #입력변수 home_search -> 소문자변환
    #set_sns_column 검색
    #if 해당 homepage 없다면 add request
    
    home_search = input("원하는 SNS및 플랫폼을 입력하세요: ")
    home_search = home_search.lower()
    
    
    for element in set_sns_column :
        if home_search in set_sns_column:
            print("해당 SNS및 플랫폼이 존재합니다.")
            return main_work()
                
        else :
            print("해당 SNS및 플랫폼이 존재하지 않습니다.")
            
            home_add = int(input("추가하시겠습니까? [yes: 1, no: 0]: "))
            if home_add == 1:
                print("시작화면에서 메뉴 2번을 눌러주세요.")
                return main_work()
            elif home_add == 0:
                print("시작 화면으로 돌아갑니다.")
                return main_work()
            else :
                print("유효한 동작이 아닙니다.")
                return main_work()
                    





def home_add():
    #SNS, 플랫폼 add function
    #SNS, 플랫폼 명 추가, 소문자 변환
    home_add = input("추가할 SNS및 플랫폼을 입력하세요: ")
    home_add = home_add.lower()
    
    #insert_rows(행 번호, 행 개수) function: 행 번호부터 행 개수만큼 새로운 행을 추가하는 함수
    #openpyxl 라이브러리 함수이다.
    sheet.insert_rows(2, 1)
    
    #추가한 행의 첫번째 셀 index = A2 => sheet[index명] 으로 데이터를 넣을 수 있다.
    sheet['A2'] = home_add
    
    #id_add() function 참고
    id_add()
    
    #save(파일경로) function -> 엑셀파일을 덮어 저장하는 함수. 엑셀파일이 열려있을 경우 error message
    return main_work()
    





def home_delete():
    #홈페이지 삭제 및 해당 행의 데이터 전체 소멸
    home_delete = input("삭제할 SNS및 플랫폼을 입력하세요: ")
    # id_delete = input("삭제할 id를 입력하세요: ")
    
    #SNS, 플랫폼 명 소문자변환 -> 검색에 용이
    home_delete = home_delete.lower()

    
    def delete(home) :
        #1행을 제외한 모든 행을 순환하며 for문 동작
        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row):
            if row[0].value == home:            #1열(SNS)의 값 = input한 SNS, 플랫폼 명과 동일
                # if row[1].value == id_delete:   #2열(ID)의 값 = input한 id와 동일
                #     #delete_rows(행번호, 행 개수) function : 행 번호부터 행 개수만큼 삭제하는 함수
                #     #openpyxl 라이브러리 함수이다. delete 시 엑셀파일 내에서는 빈 셀을 삭제하여 자동 정렬한다.
                sheet.delete_rows(row[0].row)
            #작업 후 파일 저장
            datafile.save(filepath)

    #main
    if home_delete in set_sns_column :
        print("해당 SNS및 플랫폼이 존재합니다.")
        
        #삭제요청 재 확인
        ask_delete = int(input("정말 삭제하시겠습니까?[yes: 1, no: 0]: "))
        if ask_delete == 1: 
            delete(home_delete)
            print("정상적으로 삭제되었습니다.")
            return main_work()
        
        elif ask_delete == 0:
            print("삭제 취소되었습니다. 메인 화면으로 돌아갑니다.")
            return main_work()
        
        else:
            print("유효한 요청이 아닙니다.")
            return main_work()
        
    #검색한 SNS가 존재하지 않을 경우
    else :
        print("해당 SNS및 플랫폼이 존재하지 않습니다.")
        return main_work()
            
    




def pw_search(home):
    #해당 홈페이지 내 존재하는 아이디 all show
    #아이디 검색 -> 비밀번호 show
    #if 입력 아이디가 없는 아이디 -> id add request
    id_list = []
    pw_list = []
    
    def show_id(home):
        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row):
            if row[0].value == home:
                id_list.append(row[1].value)
                pw_list.append(row[2].value)
        print("%s에 해당하는 계정입니다." %(home), end = ": ")
        print(id_list)
    
    
    #해당 플랫폼 내 모든 계정 및 비밀번호 표시
    def show_all(home):
        print("%s 내 전체 계정 표시" %home, end = "\n")
        for i in range(len(id_list)):
            print("ID: %s\t PW: %s" %(id_list[i], pw_list[i]), end = "\n")



    def find_pw():
        pw = ""         #비밀번호 저장 변수 초기화
        id_search = input("검색할 비밀번호의 아이디를 입력하세요: ")
        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row):
            if row[1].value == id_search:           #1열(id)의 값 = search id 값과 동일
                pw = row[2].value                   #해당 id 행의 3열(pw)값을 pw변수에 저장
        
        if pw != None:
            print("계정 %s의 비밀번호: %s" %(id_search, pw))
            
        #id, pw는 id_add함수에서 한번에 저장하는 방식이므로
        #pw값이 없을 경우 id 값도 존재하지 않음
        else:
            print("해당하는 아이디가 없습니다.")
      
    #main
    show_id(home)
    ask = int(input("전체 계정 표시는 1번을 입력해주세요: "))
    if ask == 1:
        show_all(home)
        return main_work()
    else:
        find_pw()
    return main_work()
    
    
    
    
    
    
    
#초기 제작 의도는 SNS 추가 단일동작, 아이디+비밀번호 추가 단일동작으로 두 가지의 add function을 구현하려 하였으나
#SNS와 아이디, 비밀번호의 전체 추가 방식으로 최종 제작되어
#id_add() 함수는 home_add() 함수와 연계하여 사용하는 것으로 제한한다.
#id_add() 단독 사용할 경우, 입력데이터 값이 SNS에 상관없이 2열2행, 2열3행에 자동 덮어쓰기되어
#데이터의 변질 우려가 있다.
def id_add():
#아이디, 비밀번호 동시 저장.
    id_add = input("계정 아이디 입력: ")
    pw_add = input("계정 비밀번호 입력: ")
    
    #input한 id, pw를 data변수에 리스트 형태로 저장
    data = [id_add, pw_add]
    
    #enumerate() = index를 부여 -> data list의 원소에 index '2'부터 index를 부여한다.
    #해당 for문은 column index를 부여하며, 각 value는 data list의 element에 해당한다.
    for col_num, value in enumerate(data, start=2):
        
        #즉 enumerate로 data list의 원소에 index를 할당하며, 
        #target이라는 일종의 포인터 변수에 destination 위치를 지정한다.
        target = sheet.cell(row=2, column=col_num)
        
        #target(2행, comlumn열) 위치에 value(data element) 를 저장한다.
        target.value = value
        
        #위 for문은 data list[id_add, pw_add]의 원소 개수인 2번 진행하며,
        #2행 2열에 id_add, 2행 3열에 pw_add를 저장한다.
    
    #위 동작 후 엑셀파일을 저장
    datafile.save(filepath)
    print("정상적으로 추가되었습니다.")
    
    






def id_delete(home) :
    #기본 아이디 삭제, 해당 아이디, 연결된 비밀번호 소멸
    id_list = []
    
    def show_id(home):
        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row):
            if row[0].value == home:
                id_list.append(row[1].value)
        print("%s에 해당하는 계정입니다." %(home), end = ": ")
        print(id_list)
    
    
    def delete():
        id_search = input("아이디를 입력하세요: ")
        
        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row):
            if row[1].value == id_search:
                ask = int(input("정말 삭제하시겠습니까?[yes: 1, no: 0]: "))
                if ask == 1:
                    sheet.delete_rows(row[1].row)
                    print("정상적으로 삭제되었습니다.")
                elif ask == 0:
                    print("메인화면으로 돌아갑니다.")
                    return main_work()
                else :
                    print("유효한 요청이 아닙니다.")
                    return main_work()
                datafile.save(filepath)
                
                #1회만 작동 후 종료
                break

    #main
    show_id(home)
    delete()
    return main_work()
    
    
    
    

def pw_change(home):
    #password change function
    #변경할 비밀번호의 SNS명과 id를 입력 후 index지정
    
    def change(home) :
        pw_id_change = input("변경할 비밀번호의 아이디를 입력하세요: ")
        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row):
            if row[0].value == home :
                if row[1].value == pw_id_change:
                    new_pw = input("새 비밀번호 입력: ")
                    row[2].value = new_pw  # row[2]에 새 비밀번호 할당
                    print("정상적으로 변경되었습니다.")
                    datafile.save(filepath)
                    break
                
                else:
                    print("해당 아이디가 존재하지 않습니다.")
                    return main_work()
        
        return main_work()
        
    #main
    if home in set_sns_column :
        print("해당 SNS및 플랫폼이 존재합니다.")
        change(home)
        
    else :
        print("해당 SNS및 플랫폼이 존재하지 않습니다.")
        return main_work()


def delete_all(sheet):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        sheet.delete_rows(row[0].row)
        
    datafile.save(filepath)
    print("정상적으로 삭제되었습니다.")
    return main_work()
    



############# Main Interface Code ################
#interface function
def interface():
    print("")
    print("*" *17, end= " ")
    print("메뉴", end = " ")
    print("*" *17)
    print("[1] SNS, 홈페이지 검색\t[2] SNS, 홈페이지 추가\t[3] SNS, 홈페이지 삭제")
    print("[4] 비밀번호 검색\t\t[5] 계정 삭제\t\t[6] 비밀번호 변경")
    print("[7] 모든 계정 삭제\t\t[8] 메뉴 화면 출력")
    print("*" *40)

sns_offset = offset()
set_sns_column = set(sns_offset)


#프로그램 동작 분기코드
def main_work() :
    count_work_times = 0
    
    while True :
        work_num = int(input("\n원하는 메뉴의 번호를 입력하세요: "))
        
        if work_num == 1:
            print("[1] SNS, 홈페이지 검색")
            home_search(sns_column)
            break
            
        elif work_num == 2:
            print("[2] SNS, 홈페이지 추가")
            home_add()
            offset()
            break
            
        elif work_num == 3:
            print("[3] SNS, 홈페이지 삭제")
            home_delete()
            break
            
        elif work_num == 4:
            print("[4] 비밀번호 검색")
            home = input("검색할 아이디의 SNS및 플랫폼을 입력하세요: ")
            home = home.lower()
            pw_search(home)
            break
            
        elif work_num == 5:
            print("[5] 계정 삭제")
            home = input("삭제할 아이디의 SNS및 플랫폼을 입력하세요: ")
            home = home.lower()
            id_delete(home)
            break
        
        elif work_num == 6:
            print("[6] 비밀번호 변경")
            home = input("삭제할 아이디의 SNS및 플랫폼을 입력하세요: ")
            home = home.lower()
            pw_change(home)
            break
        
        elif work_num == 7:
            print("[7] 모든 계정 삭제")
            print("[주의] 메뉴 실행 시 모든 데이터가 삭제됩니다.")
            ask = int(input("정말 삭제하시겠습니까?[yes: 1, no: 0]: "))
            if ask == 1:
                delete_all(sheet)
            elif ask == 0:
                print("삭제 취소되었습니다.")
                return main_work()
            else:
                print("유효한 동작이 아닙니다.")
                return main_work()
            
            
        elif work_num == 8:
            print("[8] 메뉴 화면")
            interface()
            
        else :
            #규정된 동작 외 요청 거절, 5회오류 시 프로그램 강제 종료
            print("유효한 동작이 아닙니다.")
            count_work_times += 1
            if count_work_times == 5:
                print("5회 오류, 프로그램을 종료합니다.")
                break

interface()
main_work()
